"""
Load the template Excel, fill only the two yellow columns per data row,
plus header cells E5, E7, R7, R8. All existing formulas are preserved.

Template column mapping (adjust if the template changes):
  - 品號 column used for SKU lookup: COL_ITEM_NO
  - 採購數量(pcs): COL_QTY_PCS
  - 到貨日: COL_ARRIVAL_DATE
"""

import io
from datetime import date
from pathlib import Path

import openpyxl

_TEMPLATE_PATH = Path(__file__).parent / "_KAO_八達網花王採購單-可複製.xlsx"

# ── Adjust these column letters to match the actual template ──────────────────
COL_ITEM_NO = "E"       # 品號（用於比對 SKU，不寫入）
COL_QTY_PCS = "K"       # 採購數量(pcs) — 黃底，程式填入
COL_ARRIVAL_DATE = "N"  # 到貨日 — 黃底，程式填入
DATA_START_ROW = 8      # 主表格第一筆資料列
# ─────────────────────────────────────────────────────────────────────────────


def generate_excel(
    sku_qty_map: dict[str, str],
    warehouse_name: str,
    warehouse_code: str,
    warehouse_address: str,
    purchase_date: date,
    arrival_date: date,
) -> bytes:
    """
    Load template, fill header + data rows, return file bytes.
    """
    if not _TEMPLATE_PATH.exists():
        raise FileNotFoundError(
            f"找不到公版 Excel：{_TEMPLATE_PATH}\n"
            "請將 _KAO_八達網花王採購單-可複製.xlsx 放置於 stage2_purchase_order/"
        )

    wb = openpyxl.load_workbook(_TEMPLATE_PATH)
    ws = wb.active

    # ── Header cells ──────────────────────────────────────────────────────────
    ws["E5"] = warehouse_address
    ws["E7"] = purchase_date.strftime("%Y/%m/%d")
    ws["R7"] = warehouse_name
    ws["R8"] = warehouse_code

    # ── Data rows ─────────────────────────────────────────────────────────────
    row = DATA_START_ROW
    while True:
        item_no_cell = ws[f"{COL_ITEM_NO}{row}"]
        item_no = str(item_no_cell.value or "").strip()

        # Stop when we hit an empty 品號 cell (end of data block)
        if not item_no:
            break

        qty = sku_qty_map.get(item_no, "")
        if qty:
            ws[f"{COL_QTY_PCS}{row}"] = qty

        ws[f"{COL_ARRIVAL_DATE}{row}"] = arrival_date.strftime("%Y/%m/%d")

        row += 1

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
